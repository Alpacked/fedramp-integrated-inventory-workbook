# (c) 2019 Amazon Web Services, Inc. or its affiliates. All Rights Reserved.
# License:
# This sample code is made available under the MIT-0 license. See the LICENSE file.
from datetime import datetime
import logging
from pathlib import PurePath
import os, os.path
from typing import List
import boto3
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from inventory.mappers import InventoryData

from botocore.exceptions import ClientError
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

_logger = logging.getLogger("inventory.reports")
_logger.setLevel(os.environ.get("LOG_LEVEL", logging.INFO))
_current_dir_name = os.path.dirname(__file__)
_workbook_template_file_name = os.path.join(_current_dir_name, "SSP-A13-FedRAMP-Integrated-Inventory-Workbook-Template.xlsx")
_workbook_output_file_path = PurePath("/tmp/SSP-A13-FedRAMP-Integrated-Inventory.xlsx")
DEFAULT_REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER = 6

class CreateReportCommandHandler():
    def _write_cell_if_value_provided(self, worksheet: Worksheet, column:int, row: int, value: str):
        if value:
            worksheet.cell(column=column, row=row, value=value)

    def execute(self, inventory: List[InventoryData]) -> str:
        workbook = load_workbook(_workbook_template_file_name)
        reportWorksheetName = os.environ.get("REPORT_WORKSHEET_NAME", "Inventory")
        reportWorksheet = workbook[reportWorksheetName]
        rowNumber: int = int(os.environ.get("REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER", DEFAULT_REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER))

        _logger.info(f"writing {len(inventory)} rows into worksheet {reportWorksheetName} starting at row {rowNumber}")

        for inventory_row in inventory:
            self._write_cell_if_value_provided(reportWorksheet, 1, rowNumber, inventory_row.unique_id)
            self._write_cell_if_value_provided(reportWorksheet, 2, rowNumber, inventory_row.ip_address)
            self._write_cell_if_value_provided(reportWorksheet, 3, rowNumber, inventory_row.is_virtual)
            self._write_cell_if_value_provided(reportWorksheet, 4, rowNumber, inventory_row.is_public)
            self._write_cell_if_value_provided(reportWorksheet, 5, rowNumber, inventory_row.dns_name)
            self._write_cell_if_value_provided(reportWorksheet, 7, rowNumber, inventory_row.mac_address)
            self._write_cell_if_value_provided(reportWorksheet, 8, rowNumber, inventory_row.authenticated_scan_planned)
            self._write_cell_if_value_provided(reportWorksheet, 9, rowNumber, inventory_row.baseline_config)
            self._write_cell_if_value_provided(reportWorksheet, 12, rowNumber, inventory_row.asset_type)
            self._write_cell_if_value_provided(reportWorksheet, 13, rowNumber, inventory_row.hardware_model)
            self._write_cell_if_value_provided(reportWorksheet, 15, rowNumber, inventory_row.software_vendor)
            self._write_cell_if_value_provided(reportWorksheet, 16, rowNumber, inventory_row.software_product_name)
            self._write_cell_if_value_provided(reportWorksheet, 21, rowNumber, inventory_row.network_id)
            self._write_cell_if_value_provided(reportWorksheet, 22, rowNumber, inventory_row.owner)
            self._write_cell_if_value_provided(reportWorksheet, 23, rowNumber, inventory_row.owner)

            rowNumber += 1

        workbook.save(_workbook_output_file_path)

        _logger.info(f"completed saving inventory into {_workbook_output_file_path}")

        return str(_workbook_output_file_path)

class DeliverReportCommandHandler():
    def __init__(self, s3_client=boto3.client('s3')):
        self._s3_client = s3_client

    def execute(self, report_file_name: str) -> str:
        target_path = os.environ["REPORT_TARGET_BUCKET_PATH"]
        target_bucket = os.environ["REPORT_TARGET_BUCKET_NAME"]
        report_s3_key = os.path.join(target_path, f"{_workbook_output_file_path.stem}-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx")

        _logger.info(f"uploading file '{report_file_name}' to bucket '{target_bucket}' with key '{report_s3_key}'")

        object_data = open(report_file_name, "rb")

        self._s3_client.put_object(Bucket=target_bucket, Key=report_s3_key, Body=object_data)

        _logger.info(f"completed file upload")

        target_bucket_location = self._s3_client.get_bucket_location(Bucket=target_bucket)['LocationConstraint']

        return f"https://s3-{target_bucket_location}.amazonaws.com/{target_bucket}/{report_s3_key}"


def send_email_with_attachment(file_path):
    # Amazon SES Documentation
    # https://docs.aws.amazon.com/ses/latest/dg/send-email-raw.html#send-email-raw-api

    from_email = os.environ.get("SENDER")
    from_email_friendly_name = os.environ.get(
        "SENDER_FRIENDLY_NAME", 'SRE Reports')
    to_email = os.environ.get("RECIPIENT")

    SENDER = f"{from_email_friendly_name} <{from_email}>"
    RECIPIENT = to_email

    # The subject line for the email.
    SUBJECT = "FedRamp Inventory"
    # The full path to the file that will be attached to the email.
    ATTACHMENT = file_path
    # The email body for recipients with non-HTML email clients.
    DEFAULT_MESSAGE = "Please see the attached file for a FedRamp inventory report."
    BODY_TEXT = os.environ.get(
        "BODY_TEXT",
        f"Hello,\r\n{DEFAULT_MESSAGE}"
    )

    # The HTML body of the email.
    BODY_HTML = os.environ.get("BODY_HTML", f"""\
    <html>
    <head></head>
    <body>
    <h1>Hello!</h1>
    <p>{os.environ.get('BODY_HTML', DEFAULT_MESSAGE)}</p>
    </body>
    </html>
    """)

    # The character encoding for the email.
    CHARSET = "utf-8"

    # Create a new SES resource and specify a region.
    client = boto3.client('ses')

    # Create a multipart/mixed parent container.
    msg = MIMEMultipart('mixed')
    # Add subject, from and to lines.
    msg['Subject'] = SUBJECT
    msg['From'] = SENDER
    msg['To'] = RECIPIENT

    # Create a multipart/alternative child container.
    msg_body = MIMEMultipart('alternative')

    # Encode the text and HTML content and set the character encoding. This step is
    # necessary if you're sending a message with characters outside the ASCII range.
    textpart = MIMEText(BODY_TEXT.encode(CHARSET), 'plain', CHARSET)
    htmlpart = MIMEText(BODY_HTML.encode(CHARSET), 'html', CHARSET)

    # Add the text and HTML parts to the child container.
    msg_body.attach(textpart)
    msg_body.attach(htmlpart)

    # Define the attachment part and encode it using MIMEApplication.
    att = MIMEApplication(open(ATTACHMENT, 'rb').read())

    # Add a header to tell the email client to treat this part as an attachment,
    # and to give the attachment a name.
    att.add_header(
        'Content-Disposition',
        'attachment',
        filename=os.path.basename(ATTACHMENT)
    )

    # Attach the multipart/alternative child container to the multipart/mixed
    # parent container.
    msg.attach(msg_body)

    # Add the attachment to the parent container.
    msg.attach(att)
    # print(msg)
    try:
        # Provide the contents of the email.
        response = client.send_raw_email(
            Source=SENDER,
            Destinations=[
                RECIPIENT
            ],
            RawMessage={
                'Data': msg.as_string(),
            },
            # ConfigurationSetName=CONFIGURATION_SET
        )
    # Display an error if something goes wrong.
    except ClientError as e:
        print(e.response['Error']['Message'])
    else:
        print("Email sent! Message ID:"),
        print(response['MessageId'])
